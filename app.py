import sys
import os
from pathlib import Path
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog, QLabel,
    QTableWidget, QTableWidgetItem, QHeaderView, QProgressBar, QCheckBox, QHBoxLayout, QGridLayout,
    QInputDialog  
)
from PyQt6.QtGui import QPixmap, QPainter, QColor, QPen
from PyQt6.QtCore import Qt, QRectF, QPoint
from PyQt6.QtCore import pyqtSignal
from ultralytics import YOLO 
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
MODEL_PATH = BASE_DIR / "models" / "pcb_defect_yolov8_40epochs" / "weights" / "best.pt"
OUTPUT_DIR = BASE_DIR / "outputs" / "app_predictions"

DEFECT_NAMES = {
    0: "mouse_bite",
    1: "spur",
    2: "missing_hole",
    3: "short",
    4: "open_circuit",
    5: "spurious_copper"
}

class DefectViewer(QWidget):
    closed = pyqtSignal()  
    def __init__(self, image_path, defects):
        super().__init__()
        self.setWindowTitle("Просмотр дефекта")
        self.setGeometry(150, 150, 1400, 900)

        self.image_path = image_path
        self.defects = defects 
        self.selected_defect_idx = None

        self.results_data = {}

        layout = QHBoxLayout()
        self.setLayout(layout)

        self.image_label = QLabel(self)
        pixmap = QPixmap(image_path)
        self.original_pixmap = pixmap
        self.displayed_pixmap = pixmap.scaled(900, 900, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation)
        self.image_label.setPixmap(self.displayed_pixmap)
        self.image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.image_label)

        self.table_widget = QTableWidget()
        self.table_widget.setRowCount(len(defects))
        self.table_widget.setColumnCount(2)  # Название дефекта + чекбокс
        self.table_widget.setHorizontalHeaderLabels(["Дефект", "Обнаружен"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        for row, defect in enumerate(defects):
            class_idx = int(defect['class'])
            defect_name = DEFECT_NAMES.get(class_idx, f"unknown_{class_idx}")

            self.table_widget.setItem(row, 0, QTableWidgetItem(defect_name))
            checkbox = QCheckBox()
            checkbox.setChecked(defect.get('detected', True))
            checkbox.stateChanged.connect(lambda state, r=row: self.on_checkbox_state_changed(r, state))
            self.table_widget.setCellWidget(row, 1, checkbox)

        right_layout = QVBoxLayout()

        right_layout.addWidget(self.table_widget)

        self.add_defect_button = QPushButton("+")
        self.add_defect_button.setFixedSize(40, 40)  # Сделаем маленькую квадратную кнопку
        self.add_defect_button.clicked.connect(self.start_adding_defect)
        self.add_defect_button.setStyleSheet("font-size: 20px;")

        self.delete_defect_button = QPushButton("🗑️")  # <-- Кнопка удаления
        self.delete_defect_button.clicked.connect(self.delete_defect)

        button_layout = QHBoxLayout()
        button_layout.addStretch()  # Растяжение влево
        button_layout.addWidget(self.add_defect_button)
        button_layout.addWidget(self.delete_defect_button)  # <-- добавляем сюда кнопку удаления

        self.delete_defect_button.setFixedSize(40, 40)
        self.delete_defect_button.setStyleSheet("font-size: 20px;")

        right_layout.addLayout(button_layout)

        layout.addLayout(right_layout)  # Добавляем правый вертикальный лэйаут

        self.image_label.installEventFilter(self)  # Для обработки кликов по картинке
        self.table_widget.cellClicked.connect(self.on_table_click)  # Обработчик клика по таблице

        self.draw_boxes()

        self.adding_defect = False  # Флаг режима добавления

        self.current_rect = None
        self.start_point = None
        self.end_point = None
        self.draw_boxes()  # После добавления нового бокса перерисовать начисто

        self.new_defects = []

    def on_checkbox_state_changed(self, row, state):
        if 0 <= row < len(self.defects):
            self.defects[row]['detected'] = (state == Qt.CheckState.Checked)

    def closeEvent(self, event):
        self.closed.emit()
        super().closeEvent(event)

    def get_defect_states(self):
        states = []
        for row in range(self.table_widget.rowCount()):
            checkbox = self.table_widget.cellWidget(row, 1)
            if checkbox is not None:
                states.append(checkbox.isChecked())
            else:
                states.append(True)  # если почему-то чекбокса нет — считать что True
        return states

    def update_selection_rectangle(self):
        if self.start_point and self.end_point:
            scaled_pixmap = self.original_pixmap.scaled(
                900, 900, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation
            )
            painter = QPainter(scaled_pixmap)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing)

            pixmap_width = scaled_pixmap.width()
            pixmap_height = scaled_pixmap.height()
            orig_width = self.original_pixmap.width()
            orig_height = self.original_pixmap.height()

            scale_x = pixmap_width / orig_width
            scale_y = pixmap_height / orig_height

            for idx, defect in enumerate(self.defects):
                x1, y1, x2, y2 = defect['box']
                rect = QRectF(x1 * scale_x, y1 * scale_y, (x2 - x1) * scale_x, (y2 - y1) * scale_y)
                if idx == self.selected_defect_idx:
                    pen = QPen(QColor("red"), 3)
                else:
                    pen = QPen(QColor("cyan"), 1)
                painter.setPen(pen)
                painter.drawRect(rect)

            label_size = self.image_label.size()
            pixmap_size = self.displayed_pixmap.size()
            offset_x = (label_size.width() - pixmap_size.width()) / 2
            offset_y = (label_size.height() - pixmap_size.height()) / 2

            adjusted_start = QPoint(int(self.start_point.x() - offset_x), int(self.start_point.y() - offset_y))
            adjusted_end = QPoint(int(self.end_point.x() - offset_x), int(self.end_point.y() - offset_y))

            rect = QRectF(
                adjusted_start.x(), adjusted_start.y(),
                adjusted_end.x() - adjusted_start.x(), adjusted_end.y() - adjusted_start.y()
            )

            painter.setPen(QPen(QColor(0, 255, 0, 150), 2))  # Полупрозрачный зелёный контур
            painter.setBrush(QColor(0, 255, 0, 50))           # Полупрозрачная зелёная заливка
            painter.drawRect(rect)

            painter.end()

            self.displayed_pixmap = scaled_pixmap
            self.image_label.setPixmap(self.displayed_pixmap)

    def eventFilter(self, source, event):
        if source is self.image_label:
            if self.adding_defect:
                if event.type() == event.Type.MouseButtonPress:
                    self.start_point = event.position().toPoint()
                    return True
                elif event.type() == event.Type.MouseMove:
                    self.end_point = event.position().toPoint()
                    self.update_selection_rectangle()
                    return True
                elif event.type() == event.Type.MouseButtonRelease:
                    self.end_point = event.position().toPoint()
                    self.finish_adding_defect()  # <-- вот это добавить!!
                    return True
            elif event.type() == event.Type.MouseButtonPress and not self.adding_defect:
                pos = event.position().toPoint()
                clicked_idx = self.detect_clicked_box(pos)
                if clicked_idx is not None:
                    self.highlight_defect(clicked_idx)
                return True
        return super().eventFilter(source, event)

    def detect_clicked_box(self, pos):
        label_size = self.image_label.size()
        pixmap_size = self.displayed_pixmap.size()

        offset_x = (label_size.width() - pixmap_size.width()) / 2
        offset_y = (label_size.height() - pixmap_size.height()) / 2

        adjusted_x = (pos.x() - offset_x)
        adjusted_y = (pos.y() - offset_y)

        if adjusted_x < 0 or adjusted_y < 0 or adjusted_x > pixmap_size.width() or adjusted_y > pixmap_size.height():
            return None  # клик вне изображения

        scale_x = self.original_pixmap.width() / pixmap_size.width()
        scale_y = self.original_pixmap.height() / pixmap_size.height()

        clicked_x = adjusted_x * scale_x
        clicked_y = adjusted_y * scale_y

        for idx, defect in enumerate(self.defects):
            x1, y1, x2, y2 = defect['box']
            if x1 <= clicked_x <= x2 and y1 <= clicked_y <= y2:
                return idx
        return None

    def highlight_defect(self, idx):
        self.selected_defect_idx = idx
        for row in range(self.table_widget.rowCount()):
            item = self.table_widget.item(row, 0)
            if row == idx:
                item.setBackground(QColor("yellow"))
            else:
                item.setBackground(Qt.GlobalColor.transparent)
        self.draw_boxes()

    def on_table_click(self, row, column):
        self.highlight_defect(row)

    def draw_boxes(self):
        if self.original_pixmap.isNull():
            return

        scaled_pixmap = self.original_pixmap.scaled(
            900, 900, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation
        )
        painter = QPainter(scaled_pixmap)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        pixmap_width = scaled_pixmap.width()
        pixmap_height = scaled_pixmap.height()
        orig_width = self.original_pixmap.width()
        orig_height = self.original_pixmap.height()

        scale_x = pixmap_width / orig_width
        scale_y = pixmap_height / orig_height

        for idx, defect in enumerate(self.defects):
            x1, y1, x2, y2 = defect['box']
            rect = QRectF(x1 * scale_x, y1 * scale_y, (x2 - x1) * scale_x, (y2 - y1) * scale_y)

            if idx == self.selected_defect_idx:
                pen = QPen(QColor("red"), 3)
            else:
                pen = QPen(QColor("cyan"), 1)

            painter.setPen(pen)
            painter.drawRect(rect)

        painter.end()

        self.displayed_pixmap = scaled_pixmap
        self.image_label.setPixmap(self.displayed_pixmap)
    
    def start_adding_defect(self):
        self.adding_defect = True
        self.start_point = None
        self.end_point = None
        self.image_label.setCursor(Qt.CursorShape.CrossCursor)

    def finish_adding_defect(self):
        if self.start_point and self.end_point:
            label_size = self.image_label.size()
            pixmap_size = self.displayed_pixmap.size()

            offset_x = (label_size.width() - pixmap_size.width()) / 2
            offset_y = (label_size.height() - pixmap_size.height()) / 2

            adjusted_start = QPoint(int(self.start_point.x() - offset_x), int(self.start_point.y() - offset_y))
            adjusted_end = QPoint(int(self.end_point.x() - offset_x), int(self.end_point.y() - offset_y))

            if (adjusted_start.x() < 0 or adjusted_start.y() < 0 or
                adjusted_end.x() < 0 or adjusted_end.y() < 0 or
                adjusted_start.x() > pixmap_size.width() or adjusted_end.x() > pixmap_size.width() or
                adjusted_start.y() > pixmap_size.height() or adjusted_end.y() > pixmap_size.height()):
                self.adding_defect = False
                self.image_label.setCursor(Qt.CursorShape.ArrowCursor)
                return

            scale_x = self.original_pixmap.width() / pixmap_size.width()
            scale_y = self.original_pixmap.height() / pixmap_size.height()

            x1 = adjusted_start.x() * scale_x
            y1 = adjusted_start.y() * scale_y
            x2 = adjusted_end.x() * scale_x
            y2 = adjusted_end.y() * scale_y

            defect_list = list(DEFECT_NAMES.values())
            defect_name, ok = QInputDialog.getItem(self, "Выбрать дефект", "Тип дефекта:", defect_list, editable=False)

            if ok and defect_name:
                class_idx = {v: k for k, v in DEFECT_NAMES.items()}[defect_name]

                new_defect = {
                    'box': (x1, y1, x2, y2),
                    'class': class_idx,
                    'detected': False,
                    'manual': True   # <-- новый флаг
                }
                self.defects.append(new_defect)

                # image_file = os.path.basename(self.image_path)
                # if image_file not in self.results_data:
                #     self.results_data[image_file] = []
                # self.results_data[image_file].append(new_defect)

                row = self.table_widget.rowCount()
                self.table_widget.insertRow(row)
                self.table_widget.setItem(row, 0, QTableWidgetItem(defect_name))
                checkbox = QCheckBox()
                checkbox.setChecked(False)  # !!! чекбокс снят для новых дефектов
                self.table_widget.setCellWidget(row, 1, checkbox)

            self.draw_boxes()

        self.adding_defect = False
        self.image_label.setCursor(Qt.CursorShape.ArrowCursor)
        self.start_point = None
        self.end_point = None

    def get_defect_states(self):
        defect_states = []
        for row in range(self.table_widget.rowCount()):
            checkbox = self.table_widget.cellWidget(row, 1)
            is_checked = checkbox.isChecked() if checkbox else True
            defect_states.append(is_checked)
        return defect_states
    
    def delete_defect(self, index):
        if 0 <= index < len(self.defects):
            self.defects[index]['detected'] = False

        items = [f"Дефект {i+1}: класс {d['class']}" for i, d in enumerate(self.defects)]
        
        item, ok = QInputDialog.getItem(self, "Удалить дефект", "Выберите дефект для удаления:", items, 0, False)
        
        if ok and item:
            idx = items.index(item)
            if 0 <= idx < len(self.defects):
                del self.defects[idx]
                self.update_view()

    def update_view(self):
        # Объединяем старые дефекты и новые
        all_defects = self.defects + self.new_defects

        self.table_widget.setRowCount(len(all_defects))

        for row, defect in enumerate(all_defects):
            class_idx = int(defect['class'])
            defect_name = DEFECT_NAMES.get(class_idx, f"unknown_{class_idx}")

            item = QTableWidgetItem(defect_name)
            self.table_widget.setItem(row, 0, item)

            checkbox = QCheckBox()
            checkbox.setChecked(defect.get('detected', True))
            checkbox.stateChanged.connect(lambda state, r=row: self.on_checkbox_state_changed(r, state))
            self.table_widget.setCellWidget(row, 1, checkbox)

        self.selected_defect_idx = None
        self.draw_boxes()

class DefectDetectionApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Анализ дефектов")
        self.setGeometry(100, 100, 1200, 800)

        self.model = YOLO(str(MODEL_PATH))

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.select_button = QPushButton("Выбрать папку с изображениями")
        self.select_button.clicked.connect(self.select_folder)
        layout.addWidget(self.select_button)

        self.start_button = QPushButton("Начать анализ")
        self.start_button.setEnabled(False)
        self.start_button.clicked.connect(self.process_images)
        layout.addWidget(self.start_button)

        self.save_base_button = QPushButton("Сохранить базу (картинки + разметка)")
        self.save_base_button.setEnabled(False)
        self.save_base_button.clicked.connect(self.save_labels_and_images)
        layout.addWidget(self.save_base_button)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setTextVisible(False)
        layout.addWidget(self.progress_bar)

        self.progress_label = QLabel("0%")
        self.progress_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(self.progress_label)

        self.table_widget = QTableWidget()
        layout.addWidget(self.table_widget)

        self.table_widget.cellDoubleClicked.connect(self.open_selected_image)

        self.folder = ""
        self.output_folder = str(OUTPUT_DIR)
        self.results_data = {}

    def update_defects_from_viewer(self, image_file):
        viewer = getattr(self, 'viewer', None)
        if viewer:
            states = viewer.get_defect_states()
            defects = self.results_data.get(image_file, [])
            for idx, defect in enumerate(defects):
                if idx < len(states):
                    defect['detected'] = states[idx]
        self.update_main_table()

    def open_selected_image(self, row, column):
        image_file = self.table_widget.item(row, 0).text()
        image_path = os.path.join(self.output_folder, image_file)
        defects = self.results_data.get(image_file, [])

        self.viewer = DefectViewer(image_path, defects)
        self.viewer.results_data = self.results_data

        self.viewer.closed.connect(lambda: self.update_defects_from_viewer(image_file))

        self.viewer.show()

    def update_defect_states_from_viewer(self):
        if hasattr(self, 'viewer') and self.viewer:
            states = self.viewer.get_defect_states()
            image_file = os.path.basename(self.viewer.image_path)
            if image_file in self.results_data:
                for idx, defect in enumerate(self.results_data[image_file]):
                    defect['detected'] = states[idx] if idx < len(states) else True

    def select_folder(self):
        self.folder = QFileDialog.getExistingDirectory(self, "Выбрать папку с изображениями")
        if self.folder:
            self.start_button.setEnabled(True)

    def process_images(self):
        if not self.folder:
            return

        os.makedirs(self.output_folder, exist_ok=True)

        image_files = [f for f in os.listdir(self.folder) if f.lower().endswith(('.png', '.jpg', '.jpeg'))]
        if not image_files:
            return

        self.results_data.clear()
        total_images = len(image_files)
        self.progress_bar.setValue(0)
        self.progress_label.setText("0%")

        for idx, image_file in enumerate(image_files):
            image_path = os.path.join(self.folder, image_file)
            result = self.model(image_path)[0]  # <-- правильно, берем только первый результат!

            defects = []

            save_path = os.path.join(self.output_folder, image_file)
            result.save(filename=save_path)

            for box in result.boxes.data.tolist():
                x1, y1, x2, y2, score, class_idx = box
                defects.append({
                    'box': (x1, y1, x2, y2),
                    'class': int(class_idx),
                    'detected': True,
                    'manual': False  # <-- это автоматически найденный
                })

            if defects:
                self.results_data[image_file] = defects

            progress = int(((idx + 1) / total_images) * 100)
            self.progress_bar.setValue(progress)
            self.progress_label.setText(f"{progress}%")
            QApplication.processEvents()

        self.update_main_table()
        self.save_base_button.setEnabled(True)

    def update_main_table(self):
        self.table_widget.clear()
        self.table_widget.setRowCount(len(self.results_data))
        self.table_widget.setColumnCount(2)
        self.table_widget.setHorizontalHeaderLabels(["Имя файла", "Ошибки обнаружения?"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        for row, (image_file, defects) in enumerate(self.results_data.items()):
            self.table_widget.setItem(row, 0, QTableWidgetItem(image_file))

            has_manual_defect = any(defect.get('manual', False) for defect in defects)
            has_undetected_auto_defect = any(not defect.get('detected', True) and not defect.get('manual', False) for defect in defects)

            has_error = has_manual_defect or has_undetected_auto_defect

            error_item = QTableWidgetItem("Да" if has_error else "Нет")
            error_item.setFlags(Qt.ItemFlag.ItemIsEnabled)  # Чтобы нельзя было редактировать вручную
            self.table_widget.setItem(row, 1, error_item)

    def save_to_excel(self):
        if not self.results_data:
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "Сохранить в Excel", "", "Excel Files (*.xlsx)")
        if not save_path:
            return

        data = []
        for image_file, defects in self.results_data.items():
            row = {"Имя файла": image_file}
            defect_counter = {DEFECT_NAMES[i]: 0 for i in DEFECT_NAMES}

            for defect in defects:
                defect_name = DEFECT_NAMES.get(defect['class'], f"unknown_{defect['class']}")
                defect_counter[defect_name] += 1

            row.update(defect_counter)
            data.append(row)

        df = pd.DataFrame(data)
        df.to_excel(save_path, index=False)

    def save_labels_and_images(self):
        if not self.results_data:
            return

        viewer = getattr(self, 'viewer', None)
        if viewer:
            states = viewer.get_defect_states()
            image_file = os.path.basename(viewer.image_path)
            if image_file in self.results_data:
                for idx, defect in enumerate(self.results_data[image_file]):
                    defect['detected'] = states[idx] if idx < len(states) else True

        save_folder = QFileDialog.getExistingDirectory(self, "Выбрать папку для сохранения")
        if not save_folder:
            return

        images_output_dir = os.path.join(save_folder, "images")
        labels_output_dir = os.path.join(save_folder, "labels")

        os.makedirs(images_output_dir, exist_ok=True)
        os.makedirs(labels_output_dir, exist_ok=True)

        excel_rows = []

        for image_file, defects in self.results_data.items():
            src_path = os.path.join(self.folder, image_file)
            dst_path = os.path.join(images_output_dir, image_file)
            if os.path.exists(src_path):
                from shutil import copyfile
                copyfile(src_path, dst_path)

            label_file = os.path.splitext(image_file)[0] + ".txt"
            label_path = os.path.join(labels_output_dir, label_file)

            pixmap = QPixmap(src_path)
            img_width = pixmap.width()
            img_height = pixmap.height()

            lines = []
            defect_idx = 1  # Начинаем с 1, но увеличиваем только если сохранили детект

            for defect in defects:
                x1, y1, x2, y2 = defect['box']
                class_idx = defect['class']
                detected = defect.get('detected', True)

                if detected:
                    # Сохраняем в label только если detected = True
                    x_center = ((x1 + x2) / 2) / img_width
                    y_center = ((y1 + y2) / 2) / img_height
                    width = (x2 - x1) / img_width
                    height = (y2 - y1) / img_height

                    lines.append(f"{class_idx} {x_center:.4f} {y_center:.4f} {width:.4f} {height:.4f}")

                error_detected = "Да" if (defect.get('manual', False) or not detected) else "Нет"

                excel_rows.append({
                    "Имя файла": image_file,
                    "ID дефекта": class_idx,
                    "Название дефекта": DEFECT_NAMES.get(class_idx, f"unknown_{class_idx}"),
                    "x1": round(x1),
                    "y1": round(y1),
                    "x2": round(x2),
                    "y2": round(y2),
                    "Ошибка обнаружения?": error_detected
                })

                defect_idx += 1  # Номер дефекта в Excel всегда увеличиваем

            with open(label_path, "w") as f:
                f.write("\n".join(lines))

        if excel_rows:

            df = pd.DataFrame(excel_rows)
            excel_path = os.path.join(save_folder, "annotations.xlsx")
            df.to_excel(excel_path, index=False)

            wb = load_workbook(excel_path)
            ws = wb.active

            filename_col = 1  # "Имя файла" — это первый столбец (A)

            current_value = None
            start_row = None

            for row in range(2, ws.max_row + 2):  # начинаем с 2, чтобы пропустить заголовок
                cell_value = ws.cell(row=row, column=filename_col).value

                if cell_value != current_value:
                    if start_row and row - start_row > 1:
                        ws.merge_cells(
                            start_row=start_row,
                            start_column=filename_col,
                            end_row=row - 1,
                            end_column=filename_col
                        )
                        cell = ws.cell(start_row, filename_col)
                        cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')

                    current_value = cell_value
                    start_row = row

            wb.save(excel_path)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = DefectDetectionApp()
    window.show()
    sys.exit(app.exec())